#-*-coding:utf-8-*-
#＝＝＝＝＝＝采购定价系统＝＝＝＝＝＝＝＝
#Python 3.7
#Author: Xiang Fu
#Email: tech@zhongwenshu.de



import sys
import xml
import lxml
import lxml.html
import openpyxl.workbook
import re
import requests
import Spider
import json
import os
import unittest


def get_transports_info():
    res = {}
    jsonstring = open(".\\Globconfig.json",'rb').read()
    allinfo = json.loads(jsonstring)
    for config in allinfo["configurations"]:
        if "transport" in config:
            for trans in config["transport"]:
                res[trans["code"]] = (trans["en"],trans["cn"],trans["consignee"])
            break
    return res

def get_ddusers_info():
    res = {}
    jsonstring = open(".\\Globconfig.json",'rb').read()
    allinfo = json.loads(jsonstring)
    for config in allinfo["configurations"]:
        if "dduser" in config:
            for user in config["dduser"]:
                res[user["code"]] = user["consignee"]
            break
    return res



def get_excel_name():
    return "_books.xlsx"


class Visitor:
    def __init__(self,file,tuan):
        htmlstring = open(file,'rb').read()
        htmltree = lxml.html.document_fromstring(self.preprocess(htmlstring))
        self._htmltree = htmltree
        self._tuan = tuan

    #预处理，比如先删除一些非法字符
    def preprocess(self,htmlstring):
        try:
            res = htmlstring.decode('utf8')
        except Exception as error:
            print(error)
            print('delete illegal multibyte sequence...')
            pos = re.findall(r'decodebytesinposition(\d+)-(\d+):illegal',str(error).replace(' ',''))
            if len(pos) != 0:
                htmlstring = htmlstring[0:int(pos[0][0])] + htmlstring[int(pos[0][1]):]
                return self.preprocess(htmlstring)
        else:
            return res

    def getOriginalPrice(self,url):
        spider = Spider.Spider(url)
        return Spider.searchOriginalPrice(spider.getHtmltree())


    #订单页面新老版本区别
    #新版本每个分包一个订单页面，老版本所有分包在一个订单页面
    def searchOrderGoodsNew(self):
        basepath = '//*[@id="__layout"]//*[@class="container"]'
        
        books = self._htmltree.xpath(basepath + '//tbody[@class="ant-table-tbody"]/tr/td[2]//*[@class="product-name"]')
        titles = self._htmltree.xpath(basepath + '//tbody[@class="ant-table-tbody"]/tr/td[2]//*[@class="product-name"]//a[@class="pro-name"]')
        hrefs = self._htmltree.xpath(basepath + '//tbody[@class="ant-table-tbody"]/tr/td[2]//*[@class="product-name"]//@href')
        prices = self._htmltree.xpath(basepath + '//tbody[@class="ant-table-tbody"]/tr/td[3]')
        bonuses = self._htmltree.xpath(basepath + '//tbody[@class="ant-table-tbody"]/tr/td[6]')
        amounts = self._htmltree.xpath(basepath + '//tbody[@class="ant-table-tbody"]/tr/td[4]')
        sums = self._htmltree.xpath(basepath + '//tbody[@class="ant-table-tbody"]/tr/td[7]')

        wb = openpyxl.Workbook()
        ws = wb.active

        j = 0
        for i,book in enumerate(books):
            #商品名称
            res = book.xpath('./*[@class="pro-tag"]/text()')
            if len(res) != 0: #有标签
                ws.cell(row=i+j+1,column=1,value='[' + res[0] + ']' + titles[i].text).hyperlink = hrefs[i]
            else:
                ws.cell(row=i+j+1,column=1,value=titles[i].text).hyperlink = hrefs[i]

            ws.cell(row=i+j+1,column=2,value=prices[i].text)
            ws.cell(row=i+j+1,column=3,value=bonuses[i].text)
            ws.cell(row=i+j+1,column=4,value=amounts[i].text)

            #小计以数字形式保存
            sum = re.findall(r'\d+.\d+',sums[i].text)[0]
            ws.cell(row=i+j+1,column=5,value=sum)

            #当当编号
            sn = Spider.split_ddsn(hrefs[i])

            #ISBN
            spider = Spider.Spider(hrefs[i])
            ws.cell(row=i+j+1,column=7,value=sn)
            ws.cell(row=i+j+1,column=8,value=spider.searchISBN())

        lastrow = len(books)


        #物流信息,采购账号
        header = ""
        consignee = ""
        payment = ""
        logiscompany = ""
        logiscn = ""
        receiverinfos = self._htmltree.xpath(basepath + '//*[@class="delivery-info"]//*[@class="receiver-info"]')
        if len(receiverinfos) != 0:
            for info in receiverinfos[0].xpath('//*[@class="item__label"]'):
                if re.match(u'.*收货人.*',info.text):
                    consignee = info.xpath('./parent::div')[0].xpath('./*[@class="item__text"]')[0].text
                elif re.match(u'.*付款方式.*',info.text):
                    payment = info.xpath('./parent::div')[0].xpath('./*[@class="item__text"]')[0].text
                elif re.match(u'.*配送公司.*',info.text):
                    logiscompany = info.xpath('./parent::div')[0].xpath('./*[@class="item__text"]')[0].text
                elif re.match(u'.*包裹号.*',info.text):
                    logiscn = info.xpath('./parent::div')[0].xpath('./*[@class="item__text"]')[0].text

            for code,(en,cn,pattern) in get_transports_info().items():
                if re.match(pattern,consignee):
                    header += u"【" + cn + u"】"
                    break

            for code,pattern in get_ddusers_info().items():
                if re.match(pattern,consignee):
                    header += u"【" + code + u"】"
                    break

        #订单号，订单状态
        baseinfo = self._htmltree.xpath(basepath + '//*[@class="order-info"]//*[@class="base-info"]')
        if len(baseinfo) != 0:
            orderid = ""
            for t in baseinfo[0].xpath('//*[@class="order-id"]/span/text()'):
                orderid += t
            orderstatus = baseinfo[0].xpath('//*[@class="order-status-desc"]/text()')[0]

        #下单时间
        routeinfo = self._htmltree.xpath(basepath + '//*[@class="order-info"]//*[@class="route-info"]')
        if len(routeinfo) != 0:
            for info in routeinfo[0].xpath('//*[@class="route-list"]//*[@class="route-li__title"]'):
                if info.xpath('./span[@class="number"]')[0].text == '01':
                    ordertime = info.xpath('./span[@class="txt"]')[0].text
                    for t in info.xpath('./parent::div')[0].xpath('./*[@class="route-li__time"]/text()'):
                        ordertime += t + " "
                    break

        ws.cell(row=lastrow+1, column=1, value = header + orderid + orderstatus + ordertime + payment)

        #商品金额总计，最终价，折扣
        amountinfo = self._htmltree.xpath(basepath + '//*[@class="produc-info__amount bottom"]/*[@class="amount-item"]')
        for i,info in enumerate(amountinfo):
            if i == 0:
                sum = info.xpath('./span[@class="amount-value"]')[0].text
                ws.cell(row=lastrow+1, column=5, value=sum.replace(u'\xa5',u'')) #删除¥符号
            elif i == len(amountinfo) - 1:
                endprice = info.xpath('./span[@class="amount-value last"]')[0].text
                ws.cell(row=lastrow+1, column=6, value=endprice.replace(u'\xa5',u'')) #删除¥符号
            else:
                bonus = info.xpath('./span[@class="amount-value"]')[0].text
                ws.cell(row=lastrow+i, column=3, value=bonus)

        #快递单号
        ws.cell(row=lastrow+1,column=12,value=logiscompany+logiscn)

        wb.save(get_excel_name())




    def searchOrderGoods(self):
        #包裹内容
        #每个订单可能有若干个分包裹，每个分包可能有若干个包件
        #没有包件的包裹路径为<table class="tabl_merch">
        #有包件的包裹中每个包件的路径为<table class="tabl_merch sort_package">
        basepath = '//*[@id="normalorder"]//*[@class="merch_bord"]//table[@class="tabl_merch" or contains(@class,"sort_package")]//tr'

        books = self._htmltree.xpath(basepath + '//*[@class="tab_w1"]/*[@name="productname"]')
        titles = self._htmltree.xpath(basepath + '//*[@class="tab_w1"]/*[@name="productname"]/@title')
        hrefs = self._htmltree.xpath(basepath + '//*[@class="tab_w1"]/*[@name="productname"]/@href')
        prices = self._htmltree.xpath(basepath + '//*[@class="tab_w3"]')
        bonuses = self._htmltree.xpath(basepath + '//*[@class="tab_w2"]')
        amounts = self._htmltree.xpath(basepath + '//*[@class="tab_w6"]')
        sums = self._htmltree.xpath(basepath + '//*[@class="tab_w4"]')
        
        #换购商品或分册信息
        subbooks = self._htmltree.xpath(basepath + '//*[@class="tab_w1"]/*[@class="present"]')

        #分包路径
        orderparcel = self._htmltree.xpath('//*[@id="normalorder"]//div[@id="divorderparcelhead"]/div[@class="order_news"]')
        parcel = self._htmltree.xpath('//*[@id="normalorder"]//div[@class="business_package"]')

        #不分包路径
        order = self._htmltree.xpath('//*[@id="normalorder"]//div[@id="divorderhead"][@class="order_news"]')
        ordertime = self._htmltree.xpath('//*[@id="normalorder"]//div[@id="divorderhead"][@class="order_news"]//span[@class="order_news_hint"]/span')
        others = self._htmltree.xpath('//*[@id="normalorder"]//div[@class="ditail_frame_notop"]/table[@class="tabl_other"]')
        endprice = self._htmltree.xpath('//*[@id="normalorder"]//div[@class="price_total"]/span[1]')
        payment = self._htmltree.xpath('//*[@id="normalorder"]//*[@class="order_detail_frame"]/ul[position()=4]/li')

        #国内物流信息
        #没有包件的包裹路径为<p class="p_space">
        #有包件的包裹中每个包件的路径为<p class="p_space" id="express_detail_?_?">
        logisexprs = []
        if len(order) != 0: #普通订单不分包裹
            logisnode = order[0].xpath('./p[@class="p_space"]')
            packagenodes = order[0].xpath('./div[@class="send_package_list"]/p[@class="p_space" and contains(@id,"express_detail")]')
            if packagenodes: #有包件
                for package in packagenodes:
                    for i,logistext in enumerate(package.xpath('./text()')):
                        if re.match(u'.*公司',logistext):
                            logiscompany = package.xpath('./span')[i]
                        if re.match(u'.*包裹号',logistext):
                            logisnr = package.xpath('./span')[i]
                    if re.match('express_detail_.+_0',package.attrib['id']): #第一个包件
                        logisexprs.append(logiscompany.text + logisnr.text)
                    else:
                        logisexprs[len(logisexprs)-1] += ',' + logiscompany.text + logisnr.text
            elif logisnode: #没有包件
                logiscompany = logisnode[0].xpath('./span[4]/span')
                logisnr = logisnode[0].xpath('./span[7]/span')
                if (logiscompany):
                    logisexprs.append(logiscompany[0].text + logisnr[0].text)
                else:
                    logisexprs.append('')
            else: #未发货
                    logisexprs.append('')
        elif len(orderparcel) != 0: #分包裹
            for parcelnode in parcel:
                logisnode = parcelnode.xpath('./p[@class="p_space"]')
                packagenodes = parcelnode.xpath('./div[@class="send_package_list"]/p[@class="p_space" and contains(@id,"express_detail")]')
                if packagenodes: #分包内有包件
                    for package in packagenodes:
                        for i,logistext in enumerate(package.xpath('./text()')):
                            if re.match(u'.*公司',logistext):
                                logiscompany = package.xpath('./span')[i]
                            if re.match(u'.*包裹号',logistext):
                                logisnr = package.xpath('./span')[i]
                        if re.match('express_detail_.+_0',package.attrib['id']): #第一个包件
                            logisexprs.append(logiscompany.text + logisnr.text)
                        else:
                            logisexprs[len(logisexprs)-1] += ',' + logiscompany.text + logisnr.text
                elif logisnode: #分包内没有包件
                    logiscompany = logisnode[0].xpath('./span[4]/span')
                    logisnr = logisnode[0].xpath('./span[7]/span')
                    if (logiscompany):
                        logisexprs.append(logiscompany[0].text + logisnr[0].text)
                    else:
                        logisexprs.append('')
                else: #分包未发货
                    logisexprs.append('')

        #单独一个包裹且没有发货的情况
        if len(logisexprs) == 0:
            logisexprs.append('')

        #国际物流信息
        header = ""
        consignee = self._htmltree.xpath('//*[@id="label_name"]')[0].text
        for code,(en,cn,pattern) in get_transports_info().items():
            if re.match(pattern,consignee):
                header += u"【" + cn + u"】"
                break

        #采购账号
        for code,pattern in get_ddusers_info().items():
            if re.match(pattern,consignee):
                header += u"【" + code + u"】"
                break


        wb = openpyxl.Workbook()
        ws = wb.active
        j = 0
        separator = []
        for i,book in enumerate(books):
            #包件分隔位置
            #分包每行路径<tr>
            #包件每行路径<tr class="merch_present">
            rowbook = book.xpath('../parent::tr[@class="merch_present"]')
            if rowbook: #包件
                packnumb = rowbook[0].xpath('./td[@class="package_numb"]')
                if packnumb: #包件内第一本
                    if i != 0:
                        separator.append(i)
                        titles[i] = '[BJ]' + titles[i]
            else: #分包
                rowbook = book.xpath('../parent::tr')
                prerowbook = rowbook[0].xpath('./preceding-sibling::tr')
                if not prerowbook: #分包内第一本
                    if i != 0:
                        separator.append(i)
                        titles[i] = '[FB]' + titles[i]

            #商品名称
            res = book.xpath('../span[@class="c_red"]')
            if len(res) != 0: #是预售
                ws.cell(row=i+j+1,column=1,value='[YS] ' + titles[i]).hyperlink = hrefs[i]
            else:
                ws.cell(row=i+j+1,column=1,value=titles[i]).hyperlink = hrefs[i]

            if len(prices[i].xpath('./text()')) != 0: #没有折扣的情况，比如订单35737447378 
                ws.cell(row=i+j+1,column=2,value=prices[i].text)
            else:
                res = prices[i].xpath('./span')
                ws.cell(row=i+j+1,column=2,value=res[0].text)

            ws.cell(row=i+j+1,column=3,value=bonuses[i].text)
            ws.cell(row=i+j+1,column=4,value=amounts[i].text)

            #小计以数字形式保存
            sum = re.findall(r'\d+.\d+',sums[i].text)[0]
            ws.cell(row=i+j+1,column=5,value=sum)

            #当当编号
            sn = Spider.split_ddsn(hrefs[i])

            #团购所需图书信息
            #团购表和采购表从第7行开始有区别
            spider = Spider.Spider(hrefs[i])
            if self._tuan:    
                titlesn = ws.cell(row=i+j+1,column=1).value + ' [' + sn  + ']'
                ws.cell(row=i+j+1,column=1,value=titlesn)
                ws.cell(row=i+j+1,column=7,value=self.getOriginalPrice(hrefs[i]))
                ws.cell(row=i+j+1,column=8,value=sn)

                ws.cell(row=i+j+1,column=9,value=spider.searchISBN())
                ws.cell(row=i+j+1,column=10,value=spider.searchPress())

                adress = spider.searchSmallAndBigPicture()
                if adress:
                    ws.cell(row=i+j+1,column=11,value=adress[0])
                    ws.cell(row=i+j+1,column=12,value=adress[1])
            else:
                ws.cell(row=i+j+1,column=7,value=sn)
                ws.cell(row=i+j+1,column=8,value=spider.searchISBN())
                

            #换购商品或分册信息
            res = books[i].xpath('../br')
            subbook = books[i].xpath('../span[@class="present"]')
            for s,elem in enumerate(subbook):
                j += 1
                hgtitle = elem.xpath('../a/@title')
                hghref = elem.xpath('../a/@href')
                hgprice = prices[i].xpath('./span/text()')
                hgamount = amounts[i].xpath('./text()')
                hgsum = sums[i].xpath('./text()')
                
                stext = elem.xpath('./text()')
                if re.match(u'.*换购',stext[0]): #有换购
                    ws.cell(row=i+j+1,column=1,value='[HG] ' + hgtitle[1+s]).hyperlink = hghref[1+s]
                else: #有分册
                    ws.cell(row=i+j+1,column=1,value='[FC] ' + hgtitle[1+s]).hyperlink = hghref[1+s]

                ws.cell(row=i+j+1,column=2,value=hgprice[s])
                if amounts[i].text:
                    ws.cell(row=i+j+1,column=4,value=hgamount[1+s])
                else:
                    ws.cell(row=i+j+1,column=4,value=hgamount[s])
                ws.cell(row=i+j+1,column=5,value=re.findall(r'\d+.\d+',hgsum[1+s])[0])
                ws.cell(row=i+j+1,column=7,value=Spider.split_ddsn(hghref[1+s]))

        lastrow = len(books) + len(subbooks)


        if len(order) != 0: #普通订单不分包裹
            #订单号，下单时间，付款方式等
            nr = ''
            for n in order[0].xpath('p/text()'):
                if n.strip() != '':
                    nr = n.strip()
                    break
            if len(ordertime) == 0:
                ws.cell(row=lastrow+1,column=1,value=header+nr+payment[0].text)
            elif len(ordertime) == 1:
                ws.cell(row=lastrow+1,column=1,value=header+nr+ordertime[0].text+payment[0].text)
            elif len(ordertime) == 2:
                ws.cell(row=lastrow+1,column=1,value=header+nr+ordertime[0].text+ordertime[1].text+payment[0].text)
            #快递单号
            ws.cell(row=lastrow+1,column=12,value=logisexprs[0])
            #最终价
            if (endprice[0].text.find(u'\xa5')) >= 0: #包含¥符号
                ws.cell(row=lastrow+1,column=6,value=endprice[0].text.replace(u'\xa5',u''))
            else:
                ws.cell(row=lastrow+1,column=6,value=endprice[0].text)
            #优惠
            bonus = others[0].xpath('.//span')
            for i,elem in enumerate(bonus):
                if i == 0:
                    if (bonus[0].text.find(u'\xa5')) >= 0: #包含¥符号
                        ws.cell(row=lastrow+1,column=5,value=bonus[0].text.replace(u'\xa5',u''))
                    else:
                        ws.cell(row=lastrow+1,column=5,value=bonus[0].text)
                else:
                    ws.cell(row=lastrow+1+i-1,column=3,value=bonus[i].text)
        elif len(orderparcel) != 0: #分包裹
            for i,elem in enumerate(parcel):
                note = elem.xpath('.//span[@class="business_package_bg"]/b/text()')
                nr = elem.xpath('.//span[@class="business_package_bg"]/text()[1]')
                time = elem.xpath('.//span[@class="business_package_bg"]//span[@class="t_time_n"]')
                ws.cell(row=lastrow+1+i,column=1,value=header+note[0]+nr[0]+time[0].text+payment[0].text)
                ws.cell(row=lastrow+1+i,column=12,value=logisexprs[i])
                ws.cell(row=lastrow+1+i,column=6,value=endprice[i].text)
                bonus = others[i].xpath('.//span')
                ws.cell(row=lastrow+1+i,column=5,value=bonus[0].text)


        wb.save(get_excel_name())



def visitorStart(file,tuan=False):
    print("Starting visitor...\n")
    visitor = Visitor(file,tuan)
    visitor.searchOrderGoodsNew()
    #visitor.searchOrderGoods()
    os.system("start " + get_excel_name())
    print("\nFinished.")




class TestVisitor(unittest.TestCase):
    def setUp(self):
        self._cwd = os.getcwd()

    def tearDown(self):
        os.chdir(self._cwd)

    def testSearchOrderGoods(self):
        files = [".\\testdata\\simple.order.dangdang.html",  #无分包无包件
                 ".\\testdata\\baojian.order.dangdang.html", #无分包有包件
                 ".\\testdata\\fenbao.order.dangdang.html",  #有分包无包件
                 ".\\testdata\\shipped.order.dangdang.html",
                 ".\\testdata\\notshipped.order.dangdang.html",
                 ".\\testdata\\fenbao_notshipped.order.dangdang.html",
                 ".\\testdata\\combi.order.dangdang.html"]   #有分包有包件
        for file in files:
            succeed = False
            try:
                visitor = Visitor(file,False)
                visitor.searchOrderGoods()
            except:
                print("unexpected error: {0} at {1}".format(sys.exc_info()[0],file))
            else:
                succeed = True
            self.assertEqual(succeed,True)

    @unittest.skip("I don't want to run this case.") #skip装饰器
    def testSearchOrderGoodsTuan(self):
        files = [".\\testdata\\simple.order.dangdang.html",
                 ".\\testdata\\fenbao.order.dangdang.html"]
        for file in files:
            succeed = False
            try:
                visitor = Visitor(file,True)
                visitor.searchOrderGoods()
            except:
                print("unexpected error: {0} at {1}".format(sys.exc_info()[0],file))
            else:
                succeed = True
            self.assertEqual(succeed,True)