#-*-coding:utf-8-*-
#＝＝＝＝＝＝＝＝网页爬取＝＝＝＝＝＝＝＝
#Python 3.7
#Author: Xiang Fu
#Email: tech@zhongwenshu.de

import sys
import webbrowser
import xml.etree.ElementTree
import openpyxl.workbook

# 正则
import re
# 操作系统功能
import os
# 网络交互
import requests
# XPath
import lxml.etree
import json

import codecs
#sys.stdout = codecs.getwriter('utf8')(sys.stdout)

import utility
import ImageProcess


#辅助函数
def remove_noise(section):
    res = section
    for each in re.findall(u'<.*?>',section,re.S):
        res = res.replace(each,u'').replace(u'\r\n',u'')
    return res

#从url提取当当编号
def split_ddsn(url):
    res = re.split('/',url)
    ddsn = re.split('\.',res[len(res)-1])
    return ddsn[0]

#获取书籍信息的具体位置
def get_xpath_indexs(tree,path,attrs,separator):
    indexs = {}
    nodes = tree.xpath(path)
    for attr in attrs:
        for i,node in enumerate(nodes):
            name = re.split(separator,node.text)[0]
            if re.match(u'.*'+attr,name.replace(' ','')):
                indexs[attr] = i+1
                break
    return indexs

#爬虫类
class Spider:
    def __init__(self,url):
        # 获取网页源代码
        self._html = utility.get_html_text(url)
        parser = lxml.html.HTMLParser()
        self._htmltree = xml.etree.ElementTree.fromstring(self._html,parser)
        self._title = u""
        self._page = u""

    def getHtml(self):
        return self._html

    def getTitle(self):
        return self._title

    def getPage(self):
        return self._page

    def getHtmltree(self):
        return self._htmltree

    #找缩略图
    def searchSmallPicture(self):
        smallpics = self._htmltree.xpath('//*[@id="largePic"]/@src')
        smallpic_urls = re.findall('http://.*.jpg',smallpics[0])
        if smallpic_urls:
            webbrowser.open(smallpic_urls[0])

    #找大图
    def searchPicture(self):
        #找到第一张大图
        regX = '<img alt=\"\" src=\".*.jpg\" title=\"\" id=\"modalBigImg\">'
        elem_url = re.findall(regX,self._html,re.S)
        for each in elem_url:
            #print(each)
            pic_url = re.findall('http://.*.jpg',each,re.S)
            webbrowser.open(pic_url[0])

        #找到共有几张大图
        #问号表示非贪婪匹配
        regX = '<ul id=\"mask-small-list-slider\">.*?</ul>'
        elem_url = re.findall(regX,self._html,re.S)
        numPicture = 0
        for each in elem_url:
            pic_url = re.findall('data-imghref=\"http://.*?.jpg\"',each,re.S)
            for pic in pic_url:
                print(re.findall('http://.*.jpg',pic)[0])
                numPicture += 1
        print("Here is " + str(numPicture) + " pictures!\n")



    #找缩略图和大图
    def searchSmallAndBigPicture(self):
        picadress = []

        smallpics = self._htmltree.xpath('//*[@id="largePic"]/@src')
        smallpic_urls = re.findall('http://.*.jpg',smallpics[0])
        if smallpic_urls:
            picadress.append(smallpic_urls[0])

        regX = '<img alt=\"\" src=\".*.jpg\" title=\"\" id=\"modalBigImg\">'
        bigpic_paths = re.findall(regX,self._html,re.S)
        if bigpic_paths:
            print(bigpic_paths[0])
            bigpic_urls = re.findall('http://.*.jpg',bigpic_paths[0],re.S)
            if bigpic_urls:
                picadress.append(bigpic_urls[0])

        return picadress

    #书名
    def searchTitle(self):
        title = ''
        titlenode = self._htmltree.xpath('//*[@id="product_info"]/div[1]/h1/text()')
        if titlenode:
            for n in titlenode:
                if n.strip(' \n\r'):
                    title = n.strip(' \n\r')
                    break
        return title
    
    #当当价
    def searchPrice(self):
        price = ''
        pricenode = self._htmltree.xpath('//*[@id="dd-price"]/text()')
        if pricenode:
            for each in pricenode:
                price += each
            price = price.strip(' \n\r')
        zhenode = self._htmltree.xpath('//*[@id="dd-zhe"]/text()')
        if zhenode:
            price += zhenode[0]
        return price

    #isbn
    def searchISBN(self):
        attrnames = [u'ISBN']
        attrpath = '//*[@id="detail_describe"]/ul/li'
        attrindexs = get_xpath_indexs(self._htmltree,attrpath,attrnames,u'：')
        
        isbn = ''
        if u'ISBN' in attrindexs:
            isbnnode = self._htmltree.xpath(attrpath + '[' + str(attrindexs[u'ISBN']) + ']' + '/text()')
            if isbnnode:
                for res in re.findall('[0-9]+',isbnnode[0]):
                    isbn += res
        return isbn

    #出版社
    def searchPress(self):
        press = ''
        pressnode = self._htmltree.xpath('//*[@id="product_info"]/div[2]/span[2]/a/text()')
        if pressnode:
            press = pressnode[0]
        return press


    #找书籍信息
    def searchAttr(self):
        _title = re.findall(u'<h1 title=.*?>.*?</h1>',self._html,re.S)[0]
        print(u"title:" + remove_noise(_title))
        subtitle = re.findall(u'<h2>.*?</h2>',self._html,re.S)[0]
        print(u'subtitle:' + remove_noise(subtitle))
        author = re.findall(u'<span class=\"t1\" id=\"author\" dd_name=\"\u4f5c\u8005\".*?</span>',self._html,re.S)[0]
        print(u'author:' + remove_noise(author))
        press = re.findall(u'<span class=\"t1\" dd_name=\"\u51fa\u7248\u793e\".*?</span>',self._html,re.S)[0]
        print(u'press:' + remove_noise(press))
        presstime = re.findall(u'<span class=\"t1\">\u51fa\u7248\u65f6\u95f4.*?</span>',self._html,re.S)[0]
        print(u'press time:' + remove_noise(presstime))
        _page = re.findall(u'<li>\u9875 \u6570.*?</li>',self._html,re.S)[0]
        print(u'page:' + remove_noise(_page))
        word = re.findall(u'<li>\u5b57 \u6570.*?</li>',self._html,re.S)[0]
        print(u'word:' + remove_noise(word))
        size = re.findall(u'<li>\u5f00 \u672c.*?</li>',self._html,re.S)[0]
        print(u'size:' + remove_noise(size))
        paper = re.findall(u'<li>\u7eb8 \u5f20.*?</li>',self._html,re.S)[0]
        print(u'paper:' + remove_noise(paper))
        packing = re.findall(u'<li>\u5305 \u88c5.*?</li>',self._html,re.S)[0]
        print(u'packing' + remove_noise(packing))
        isbn = re.findall(u'<li>\u56fd\u9645\u6807\u51c6\u4e66\u53f7.*?</li>',self._html,re.S)[0]
        print(u'ISBN:' + remove_noise(isbn))
        classification = re.findall(u'<li class=\"clearfix fenlei\" dd_name=\"\u8be6\u60c5\u6240\u5c5e\u5206\u7c7b\".*?>.*?</li>',self._html,re.S)[0]
        print(u'classification:' + remove_noise(classification))
        print("")

    #找订单信息
    def searchOrder(self):
        htmltree = lxml.etree.HTML(self._html)
        print(lxml.etree.tostring(htmltree,pretty_print=True))





def spiderStart(slist):
    print("Starting spider...\n")
    for string,tag in slist.items():  
        if tag == 0:
            spider = Spider(string)
            spider.searchPicture()
            #spider.searchAttr()
        elif tag == 1:
            spider = Spider(string)
            spider.searchOrder()
        elif tag == 2:
            data = json.loads(string)
            sn = data[u'sn']
            url = "http://product.dangdang.com/" + sn + ".html"
            spider = Spider(url)
            spider.searchPicture()
    print("\nFinished.")

#当当定价
def searchOriginalPrice(htmltree):
    ori = ''
    orinode = htmltree.xpath('//*[@id="original-price"]/text()')
    if orinode:
        for n in orinode:
            if n.strip(' \n\r'):
                ori = n.strip(' \n\r')
                break
    return ori

def spider_picture(url):
    spider = Spider(url)
    spider.searchPicture()

def spider_small_picture(url):
    spider = Spider(url)
    spider.searchSmallPicture()

def spider_small_and_big_picture(url):
    spider = Spider(url)
    return spider.searchSmallAndBigPicture()

def spider_to_excel(urllist):
    print("Starting spider...\n")
    wb = openpyxl.Workbook()
    ws = wb.active

    i = j = 0
    for url,tag in urllist.items():
        spider = Spider(url)
        if tag == 0:
            titlesn = spider.searchTitle() + ' [' + split_ddsn(url)  + ']'
            ws.cell(row=i+1,column=1,value=titlesn).hyperlink = url
            ws.cell(row=i+1,column=2,value=spider.searchPrice())
            ws.cell(row=i+1,column=3,value=searchOriginalPrice(spider.getHtmltree()))
            ws.cell(row=i+1,column=4,value=split_ddsn(url))
            ws.cell(row=i+1,column=5,value=spider.searchISBN())
            ws.cell(row=i+1,column=6,value=spider.searchPress())
            adress = spider.searchSmallAndBigPicture()
            ws.cell(row=i+1,column=7,value=adress[0])
            ws.cell(row=i+1,column=8,value=adress[1])
            i += 1
        elif tag == 1:
            j += 1

    wb.save('_books.xlsx')
    print("\nbooks info saved in excel\n")
    print("Finished")