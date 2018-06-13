#-*-coding:utf-8-*-
#＝＝＝＝＝＝奶粉分析器＝＝＝＝＝＝＝＝
#Python 2.7
#Author: Xiang Fu
#Email: tech@zhongwenshu.de


import sys
import xml
import lxml
import openpyxl.workbook
import re
import requests
import json
import jsonpath


def search_matched_bracket(str,bracket):
    bpair = []
    open = bracket[0]
    close = bracket[1]
    count = 0
    for pos,c in enumerate(str):
        if c == open:
            count += 1
            if count == 1:
                bpair.append(pos)
        elif c == close:
            count -= 1
            if count == 0:
                bpair.append(pos)
        if count < 0:
            break
    return bpair


def preprocess2(htmlstring):
    try:
        res = htmlstring.decode('utf-8')
    except Exception as error:
        print error
        print 'delete illegal multibyte sequence...'
        pos = re.findall('decodebytesinposition(\d+)-(\d+):illegal',str(error).replace(' ',''))
        if len(pos) != 0:
            htmlstring = htmlstring[0:int(pos[0][0])] + htmlstring[int(pos[0][1]):]
            return preprocess2(htmlstring)
    else:
        return res


def preprocess(htmlstring):
    res = htmlstring.replace('\n','')
    return res


def load_json(jsonstring):
    try:
        res = json.loads(jsonstring)
    except Exception as error:
        print error
        pos = re.findall('Expectingobject:line(\d+)column(\d+)\(char(\d+)\)',str(error).replace(' ',''))
        if pos:
            if jsonstring[int(pos[0][2])] == '}':
                jsonstring += '}' #补齐
            elif jsonstring[int(pos[0][2])] == '"':
                jsonstring += '}' #补齐
            return load_json(jsonstring)
    else:
        return res    


def aptamil():
    print("Analysing aptamil...\n")

    test = False
    htmltext = ''

    if test:
        htmltext = open('searchres.txt','r').read()
        htmltree = lxml.html.document_fromstring(preprocess(htmltext))
    else:
        url = 'https://s.taobao.com/search'
        #payload = {'q':'德语版老友记'}
        #payload = {'q':'德国爱他美白金二段','s':'0'} #s=0为第一页，s=44为第二页，以此类推
        payload = {'q':'玻尿酸','s':'0'}
        #payload = {'q':'万宝龙 Mont blanc','s':'0'}
        #payload = {'q':'安娜柏林','s':'0'}
        #payload = {'q':'lamy 凌美','s':'0'}
        htmltext = requests.get(url,params=payload).text
        parser = lxml.html.HTMLParser()
        htmltree = xml.etree.ElementTree.fromstring(preprocess(htmltext),parser)

    #打开Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    #获取script标签里的g_page_config变量，该变量是个json字符串
    scripts = htmltree.xpath('//script')
    for script in scripts:
        if script.text:
            if re.findall('g_page_config',script.text):
                bpair = search_matched_bracket(script.text,'{}')
                jsontexts = []
                if len(bpair)/2:
                    for i in range(len(bpair)/2):
                        jsontexts.append(script.text[bpair[i*2]:bpair[i*2+1]])
                        jsondata = load_json(jsontexts[i])

                        auctionses = jsonpath.jsonpath(jsondata, '$..auctions')
                        if auctionses:
                            auctions = auctionses[0]
                            for j in range(len(auctions)):
                                title = auctions[j]['raw_title']
                                url = 'http:' + auctions[j]['detail_url']
                                price = auctions[j]['view_price']
                                sales = re.findall('\d+',auctions[j]['view_sales'])[0]
                                nick = auctions[j]['nick']
                                shop = 'http:' + auctions[j]['shopLink']
                                if auctions[j]['shopcard']['isTmall'] == True:
                                    platform = '天猫'
                                else:
                                    platform = '淘宝'

                                if j == 0:
                                    ws.cell(row=1,column=1,value='商品名')
                                    ws.cell(row=1,column=2,value='价格(rmb)')
                                    ws.cell(row=1,column=3,value='销量')
                                    ws.cell(row=1,column=4,value='店铺')
                                    ws.cell(row=1,column=5,value='平台')
                                    ws.cell(row=1,column=6,value='销售额(rmb)')

                                ws.cell(row=j+2,column=1,value=title).hyperlink = url
                                ws.cell(row=j+2,column=2,value=price)
                                ws.cell(row=j+2,column=3,value=sales)
                                ws.cell(row=j+2,column=4,value=nick).hyperlink = shop
                                ws.cell(row=j+2,column=5,value=platform)
                                ws.cell(row=j+2,column=6,value=float(price)*float(sales))
                            
                            #翻页器
                            pagers = jsonpath.jsonpath(jsondata,'$..pager')
                            #pagers = jsonpath.jsonpath(jsondata,'$..pager',result_type='IPATH')
                            #print totalpages
                            if pagers:
                                print pagers[0]

                            break
                break
    
    wb.save('_taobao.xlsx')

    print("\nFinished.")