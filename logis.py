#-*-coding:utf-8-*-
# 中德物流
#Python 3.7
#Author: Xiang Fu
#Email: tech@zhongwenshu.de

import sys
import os
import hashlib
import json
import time
import utility
import pymysql
import logging
import re
import xml
import lxml
import openpyxl.workbook
import Visitor
import unittest


def import_logis_to_sql(server,logis):
    print("Starting import logis info to database...\n")

    mysql = server["mysql"]
    connection = pymysql.connect(mysql[0],mysql[1],mysql[2],mysql[3],charset=mysql[4])

    #区分测试和主数据库
    rinterTable = ""
    logiscnTable = ""
    logisdeTable = ""
    if mysql[3] == 'zhongw_test':
        rinterTable = "zws_test_railway_inter"
        logiscnTable = "zws_test_logis_cn"
        logisdeTable = "zws_test_logis_de"
    elif mysql[3] == 'zhongwenshu_db1':
        rinterTable = "zws_railway_inter"
        logiscnTable = "zws_logis_cn"
        logisdeTable = "zws_logis_de"

    try:
        railway_sn = ""
        inter_company = ""
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())
        cn_packet = {}
        de_packet = {}

        for area,info in logis.items():
            if area == "inter":
                for k,v in info.items():
                    inter_company = k
                    railway_sn = v
            elif area == "cn":
                for k,v in info.items():
                    for sn in v:
                        cn_packet[sn] = k
            elif area == "de":
                for k,v in info.items():
                    for sn in v:
                        de_packet[sn] = k


        #先写入国际段物流信息，因为国际段主键是作为其他表的外键
        #再写入国内段和德国段物流信息，可以不分先后
        with connection.cursor() as cursor:
            sql = "INSERT INTO " + rinterTable + " (`id`, `railway_sn`, `inter_log`, `inter_time`, `inter_status`, `inter_company`) \
                  VALUES (NULL, %s, '', %s, '-1', %s);"
            cursor.execute(sql,(railway_sn,timestamp,inter_company))

            #log
            LOG_FORMAT = "%(asctime)s - %(levelname)s - %(message)s"
            logging.basicConfig(filename='mylogis.log',level=logging.DEBUG,format=LOG_FORMAT)
            logging.info("----------------------------------------------------------------") #分割线
            logging.info(sql % ("'"+railway_sn+"'","'"+timestamp+"'","'"+inter_company+"'"))

            sql = "SELECT `id` FROM " + rinterTable + " WHERE `railway_sn`=%s"
            cursor.execute(sql,railway_sn)
            rinterid = cursor.fetchone()[0]
            print("inter id : " + str(rinterid))

            for sn,company in cn_packet.items():
                sql = "INSERT INTO " + logiscnTable + " (`id`, `cn_packet_sn`, `railway_id`, `cn_log`, `cn_time`, `cn_status`, `cn_company`) \
                    VALUES (NULL, %s, %s, '', %s, '-1', %s);"
                if not company:
                    mainsn = ['']
                    subsns = [[]]
                    companycode =['']
                    notes = [[]]
                    split_token(sn,mainsn,subsns,companycode,notes)
                    company = companycode[0]
                cursor.execute(sql,(sn,rinterid,timestamp,company))
                logging.info(sql % ("'"+sn+"'",rinterid,"'"+timestamp+"'","'"+company+"'"))

            for sn,company in de_packet.items():
                sql = "INSERT INTO " + logisdeTable + " (`id`, `de_packet_sn`, `railway_id`, `de_log`, `de_time`, `de_status`, `de_company`) \
                    VALUES (NULL, %s, %s, '', %s, '-1', %s);"
                cursor.execute(sql,(sn,rinterid,timestamp,company))
                logging.info(sql % ("'"+sn+"'",rinterid,"'"+timestamp+"'","'"+company+"'"))

        connection.commit()
    finally:
        connection.close()

    print("Finished.")


def get_logis_labels():
    res = {}
    jsonstring = open(".\\Globconfig.json",'rb').read()
    allinfo = json.loads(jsonstring)
    for config in allinfo["configurations"]:
        if "label" in config:
            for label in config["label"]:
                res[label["code"]] = label["name"]
            break
    return res


def get_logis_companies():
    res = {}
    jsonstring = open(".\\Globconfig.json",'rb').read()
    allinfo = json.loads(jsonstring)
    for config in allinfo["configurations"]:
        if "logisCompany" in config:
            for company in config["logisCompany"]:
                res[company["name"]] = company["code"]
            break
    return res


def get_logis_company_headers():
    res = {}
    jsonstring = open(".\\Globconfig.json",'rb').read()
    allinfo = json.loads(jsonstring)
    for config in allinfo["configurations"]:
        if "logisCompanyHeader" in config:
            for company in config["logisCompanyHeader"]:
                res[company["header"]] = company["attr"]
            break
    return res


def get_logis_operators():
    res = {}
    jsonstring = open(".\\Globconfig.json",'rb').read()
    allinfo = json.loads(jsonstring)
    for config in allinfo["configurations"]:
        if "logisOperator" in config:
            for operator in config["logisOperator"]:
                res[operator["type"]] = operator["characters"]
            break
    return res


#物流表达式特殊符号
#%railway(顺丰240564577148:10kg + YT2028772443672:点读笔 + JD001122-1-1)
logisKeywordHeader = get_logis_operators()["keywordheader"] #关键字
logisSeparator = get_logis_operators()["separator"] #分隔符
logisExplanator = get_logis_operators()["explanator"] #半角或全角冒号代表说明符
logisConnector = get_logis_operators()["connector"] #母子单号连接符

#
#以下若干函数物流表达式解析，不支持括号，和lib_logis.php一致
#
def split_logis_expr(expr):
    pattern = '[' + logisSeparator + ']'
    return re.split(pattern,expr.strip())

def split_logis_expr_and_token(expr):
    res = {}
    ids = split_logis_expr(expr)
    for id in ids:
        if id.strip():
            sn = [''] #为了达到传引用的效果，这里把字符串包裹在数组里
            subsns = [[]]
            company =['']
            notes = [[]]
            split_token(id.strip(),sn,subsns,company,notes)
            res[sn[0]+build_subsn(subsns[0])] = (company[0],notes[0])
    return res


def split_token(token,sn,subsns,company,notes):
    pattern = '[' + logisExplanator + ']'
    if re.search(pattern,token.strip()):
        tokens = re.split(pattern,token.strip())
        token = tokens[0]
        notes[0] = tokens[1:]

    pattern = '([a-zA-Z0-9' + logisConnector + ']+)'
    matches = re.split(pattern,token.strip())
    matches = list(filter(None,matches)) #去除空字符串
    if len(matches) == 1:
        matches[0] = [matches[0]]
        get_main_sn(matches[0],subsns)
        sn[0] = matches[0][0]
    else:
        matches[1] = [matches[1]]
        get_main_sn(matches[1],subsns)
        sn[0] = matches[1][0]
        company[0] = get_company_code(matches[0])

    if not company[0]:
        company[0] = split_company_header_code(sn)


def get_main_sn(sn,subsns):
    pattern = '[' + logisConnector + ']'
    if re.search(pattern,sn[0].strip()):
        sns = re.split(pattern,sn[0].strip())
        sn[0] = sns[0]
        subsns[0] = sns[1:]


def get_company_code(company):
    code = ""
    for n,c in get_logis_companies().items():
        if re.search(n,company):
            code = c
            break
    return code


def split_company_header_code(sn):
    company = ""
    hsize = []
    for h,attr in get_logis_company_headers().items():
        hsize.append(len(h))
    hsize = list(set(hsize))

    for size in hsize:
        header = ""
        if (len(sn[0]) > size):
            header = sn[0][0:size]
            code = get_company_header_code(header)
            if code:
                company = code[0]
                if code[1] == 0:
                    sn[0] = sn[0][size:]
                break
    return company


def get_company_header_code(header):
    res = []
    for h,attr in get_logis_company_headers().items():
        if (h.upper() == header.upper()):
            res = attr
            break
    return res


def build_subsn(subsns):
    res = ""
    logisConnector
    for subsn in subsns:
        res += logisConnector + subsn.strip()
    return res
#
#以上若干函数物和lib_logis.php一致
#


#从html生成物流表达式
#html
#<p><span style="color:#330099">邮政232324452-1<span class="da">到德仓</span> + JT7987979476461</span></p>
#<p><span style="color:#330099">邮政232324452-2:15kg + JT7987979476462:16kg:文具<span class="da">到德仓</span></span></p>
#nodes
#['邮政232324452-1', '到德仓', ' + JT7987979476461', '邮政232324452-2:15kg + JT7987979476462:16kg:文具', '到德仓']
#tokens
#[['邮政232324452-1'], 'da', ['', '', 'JT7987979476461'], ['邮政232324452-2:15kg ', '', '', 'JT7987979476462:16kg:文具'], 'da']
#multitokens
#{'邮政232324452-1':['da'], 'JT7987979476461':[], '邮政232324452-2:15kg':[], 'JT7987979476462:16kg:文具':['da']}
#logis expression
#%r(邮政232324452-1:da + JT7987979476461 + 邮政232324452-2:15kg + JT7987979476462:16kg:文具:da)
def generate_logis_expression_from_html(htmltree,filter=[],inverse=False):
    expression = ""
    for code,(en,cn,pattern) in Visitor.get_transports_info().items():
        nodes = htmltree.xpath('//div[@id="' + code + '" or @id="' + en + '"]/div[@class="descrip"]//span/text()')

        tokens = []
        for node in nodes:
            if node.strip():
                #物流标签属于上一个span里的单号，不能单独出现
                #注意这里是lxml模块的bug,getparent()判断错误，所以要额外检查node[0]不是分隔符
                if node.getparent().get("class") and not node.strip()[0] in logisSeparator:
                    tokens.append(node.getparent().get("class").strip())
                else:
                    #物流标签会打破原来的表达式，出现以分隔符比如逗号开头的表达式
                    tokens.append(split_logis_expr(node.strip()))

        multitokens = {}
        for i,token in enumerate(tokens):
            if isinstance(token,list):
                for t in token:
                    if t.strip():
                        multitokens[t.strip()] = []
            else: #是标签
                lastkey = list(multitokens.keys())[len(multitokens)-1]
                multitokens[lastkey].append(token)

        if filter:
            for token,labels in multitokens.items():
                found = False
                for label in labels:
                    if label.lower() in [elem.lower() for elem in filter]:
                        found = True
                        break
                #忽略没找到Filter中标签的单号，如果设置了inverse则正好相反
                if inverse:
                    if found:
                        multitokens[token] = ['ignore']
                else:
                    if not found:
                        multitokens[token] = ['ignore']
            multitokens = {key:multitokens[key] for key in multitokens if multitokens[key]!=['ignore']}

        for i,(token,labels) in enumerate(multitokens.items()):
            if i == 0:
                if expression:
                    expression += " " + logisSeparator[0] + " "
                expression += logisKeywordHeader + code + "("
            expression += token
            if labels:
                for label in labels:
                    expression += logisExplanator[0] + label
            if i < len(multitokens) - 1:
                expression += " " + logisSeparator[0] + " "
            else:
                expression += ")"
    return expression



def has_special_label(notes):
    if notes:
        for note in notes:
            for code,name in get_logis_labels().items():
                if note.lower() == code.lower():
                    return True
    return False


def get_customer_forecast(htmltree):
    forecast = []
    for code,(en,cn,pattern) in Visitor.get_transports_info().items():
        #当前只支持铁路
        if code == 'r':
            nodes = htmltree.xpath('//div[@id="' + code + '" or @id="' + en + '"]/div[@class="descrip"]//span/text()')
            for i,node in enumerate(nodes):
                if node.strip():
                    #物流标签属于上一个span里的单号，不能单独出现
                    label = node.getparent().get("class")
                    if label:
                        note = [label.strip()]
                        if has_special_label(note) and not node.strip()[0] in logisSeparator:
                            #注意这里是lxml模块的bug,getparent()判断错误，所以要额外检查node[0]不是分隔符
                            dicitems = forecast[len(forecast)-1].values()
                            #标签总是加在最后一个单号后面
                            (company,notes) = list(dicitems)[len(dicitems)-1]
                            notes.append(label.strip())
                            continue
                    forecast.append(split_logis_expr_and_token(node))
    return forecast


def generate_manifest_expression(data):
    expression = ''
    for i,(sn,company,notes) in enumerate(data):
        expression += get_company_name(company) + sn
        if notes:
            for note in notes:
                expression += logisExplanator[0] + note.strip()
        if i < len(data)-1:
            expression += ' ' + logisSeparator[0] + ' '
    return expression


def get_company_name(code):
    name = ""
    for n,c in get_logis_companies().items():
        if code.strip().lower() == c:
            name = n
            break
    return name


def generate_logis_expression_from_sql(server,logis):
    print("Starting generate logis expression from database...\n")

    mysql = server["mysql"]
    connection = pymysql.connect(mysql[0],mysql[1],mysql[2],mysql[3],charset=mysql[4])

    #区分测试和主数据库
    orderInfoTable = ""
    orderGoodsTable = ""
    if mysql[3] == 'zhongw_test':
        orderInfoTable = "ecs_test_order_info"
        orderGoodsTable = "ecs_test_order_goods"
        goodsTable = "ecs_test_goods"
        logiscnTable = "zws_test_logis_cn"
        logisOrderTable = "zws_test_logis_order"
    elif mysql[3] == 'zhongwenshu_db1':
        orderInfoTable = "ecs_order_info"
        orderGoodsTable = "ecs_order_goods"
        goodsTable = "ecs_goods"
        logiscnTable = "zws_logis_cn"
        logisOrderTable = "zws_logis_order"

    timestamp = time.strftime("%Y-%m-%d %H:%M:%S",time.localtime())
    template = ""
    sns = []
    regex = ""
    res = []

    for template,v in logis.items():
        for regex,orders in v.items():
            if regex == "1":
                for order in orders:
                    sns.append("^" + order + ".*$")
            elif regex == "0":
                for order in orders:
                    sns.append(order)
    try:
        with connection.cursor() as cursor:
            #log
            LOG_FORMAT = "%(asctime)s - %(levelname)s - %(message)s"
            logging.basicConfig(filename='mylogis.log',level=logging.DEBUG,format=LOG_FORMAT)
            logging.info("----------------------------------------------------------------") #分割线

            for sn in sns:
                orderinfos = {}
                if regex == "1":
                    #注意字段best_time被用于记录微信号
                    sql = "SELECT `order_id`,`order_sn`,`best_time`,`money_paid` FROM " + orderInfoTable + " WHERE `order_sn` REGEXP %s AND `order_id` \
                        in (SELECT `order_id` FROM " + orderGoodsTable + " WHERE `goods_id` = %s);"
                    cursor.execute(sql,(sn,template))
                    logging.info(sql % ("'"+sn+"'","'"+template+"'"))
                    orderinfos = cursor.fetchall()
                elif regex == "0":
                    #注意字段best_time被用于记录微信号
                    sql = "SELECT `order_id`,`order_sn`,`best_time`,`money_paid` FROM " + orderInfoTable + " WHERE `order_sn` = %s AND `order_id` \
                        in (SELECT `order_id` FROM " + orderGoodsTable + " WHERE `goods_id` = %s);"
                    cursor.execute(sql,(sn,template))
                    logging.info(sql % ("'"+sn+"'","'"+template+"'"))
                    orderinfos = cursor.fetchall()

                for (orderid,ordersn,wchat,paid) in orderinfos:
                    sql = "SELECT `goods_id` FROM " + orderGoodsTable + " WHERE order_id = %s;"
                    cursor.execute(sql,orderid)
                    logging.info(sql % ("'"+str(orderid)+"'"))
                    goodsids = cursor.fetchall()

                    validitem = {}
                    found = False
                    for goodsid in goodsids:
                        sql = "SELECT `cat_id`,`goods_name`,`goods_desc` FROM " + goodsTable + " WHERE goods_id = %s;"
                        cursor.execute(sql,goodsid[0])
                        logging.info(sql % ("'"+str(goodsid[0])+"'"))
                        (catid,goodsname,goodsdesc) = cursor.fetchone()
                        if catid == 82: #82表示订购分类
                            if not(re.match(r".*template.*",goodsname,re.IGNORECASE)): #非模板商品
                                found = True
                                validitem["ordersn"] = ordersn
                                validitem["logisgoodid"] = goodsid[0]
                                validitem["wchat"] = wchat
                                validitem["paid"] = paid

                                parser = lxml.html.HTMLParser()
                                htmltree = xml.etree.ElementTree.fromstring(goodsdesc,parser)
                                #非已发货或者收货确认的单号
                                validitem["expression"] = generate_logis_expression_from_html(htmltree,["FA","SH"],True)
                                #已发货或者收货确认的单号
                                validitem["expression2"] = generate_logis_expression_from_html(htmltree,["FA","SH"])

                                manifest = []
                                isOpenOrder = False
                                for forecast in get_customer_forecast(htmltree):
                                    for cnsn,(company,notes) in forecast.items():
                                        sql = "SELECT * FROM " + logiscnTable + " WHERE cn_packet_sn REGEXP %s;"
                                        cursor.execute(sql,cnsn)
                                        logging.info(sql % ("'"+cnsn+"'"))
                                        #未录入的而且没有特定标记的加入交接单和未结单
                                        if not cursor.fetchone():
                                            if not has_special_label(notes):
                                                manifest.append((cnsn,company,notes))
                                                isOpenOrder = True
                                        #录入的而且没有特定标记的加入未结单
                                        else:
                                            if not has_special_label(notes):
                                                isOpenOrder = True

                                #生成交接单表达式
                                validitem["manifest"] = generate_manifest_expression(manifest)

                                #更新物流订单表
                                if (isOpenOrder):
                                    validitem["open"] = 1
                                else:
                                    validitem["open"] = 0
                                sql = "SELECT * FROM " + logisOrderTable + " WHERE order_id=%s;"
                                cursor.execute(sql,orderid)
                                logging.info(sql % ("'"+str(orderid)+"'"))
                                rec = cursor.fetchone()
                                if rec:
                                    if (rec[2] != validitem["open"]):
                                        sql = "UPDATE " + logisOrderTable + " SET type=%s WHERE order_id=%s;"
                                        cursor.execute(sql,(validitem["open"],orderid))
                                        logging.info(sql % ("'"+str(validitem["open"])+"'","'"+str(orderid)+"'"))
                                else:
                                    sql = "INSERT INTO " + logisOrderTable + " (`id`, `order_id`, `type`, `from_func`, `last_update`, `reserve`) \
                                           VALUES (NULL, %s, '1', '1', NULL, '0');"
                                    cursor.execute(sql,orderid)
                                    logging.info(sql % ("'"+str(orderid)+"'"))

                                break

                    if found == True:
                        res.append(validitem)
    finally:
        connection.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    i = 0
    for item in res:
        if i == 0:
            ws.cell(row=1,column=1,value="订单号")
            ws.cell(row=1,column=2,value="国内物流单号")
            ws.cell(row=1,column=3,value="国内物流单号2")
            ws.cell(row=1,column=4,value="用户")
            ws.cell(row=1,column=5,value="支付")
            ws.cell(row=1,column=6,value="重量")
            ws.cell(row=1,column=7,value="备注")
            ws.cell(row=1,column=8,value="交接单")
            ws.cell(row=1,column=9,value="未结单")
            i = 1
        ws.cell(row=i+1,column=1,value=item["ordersn"])
        ws.cell(row=i+1,column=2,value=item["expression"])
        ws.cell(row=i+1,column=3,value=item["expression2"])
        ws.cell(row=i+1,column=4,value=item["wchat"])
        ws.cell(row=i+1,column=5,value=item["paid"])
        ws.cell(row=i+1,column=8,value=item["manifest"])
        ws.cell(row=i+1,column=9,value=item["open"])
        i += 1
    wb.save("_manifest.xlsx")
    os.system("start _manifest.xlsx")

    print("Finished.")



class TestLogis(unittest.TestCase):
    def setUp(self):
        self._cwd = os.getcwd()

    def tearDown(self):
        os.chdir(self._cwd)

    def testSplitter(self):
        sn = ['']
        subsns = [[]]
        company =['']
        notes = [[]]
        split_token("Dd112233-1-2:book:12kg",sn,subsns,company,notes)
        self.assertEqual(sn,["112233"])
        self.assertEqual(subsns[0],["1","2"])
        self.assertEqual(company,[get_logis_companies()["当当"]])
        self.assertEqual(notes[0],["book","12kg"])
        res = split_logis_expr_and_token("jd112233-1:todo ,, 中通223344 ,, DHL445566:books:fa   ups1Z")
        self.assertEqual(len(res),4)
        self.assertEqual(list(res.keys())[0],"jd112233-1")
        self.assertEqual(list(res.values())[1],(get_logis_companies()["中通"],[]))
        company,notes = list(res.values())[2]
        self.assertEqual(has_special_label(notes),True)
        self.assertEqual(list(res.keys())[3],"1Z")

    def getGoodsDesc(self):
        goodsdesc = '''
            <div><zws-product>
            <p><font size="6">2019110192875</font></p>
            <p>&nbsp;</p>
            <div class="section" id="railway">
            <div class="title"><span>铁路</span></div>
            <div class="descrip">
            <p><span style="color:#330099">邮政23232445<span class="da">到仓</span> , JT798797947646<span class="fa">德仓发出</span></span></p>
            <p><span style="color:#330099">中通232324452-1-2<span class="da">到仓</span> ,, JT7987979476461</span></p>
            <p><span style="color:#330099">中通快递232324452-2-2:玩具 ,， JT7987979476462:16kg：文具<span class="da">到仓</span>  </span></p>
            <p>各种书</p>
            <p>&nbsp;</p>
            </div>
            </div>
            </zws-product></div>
        '''
        return goodsdesc

    def testGenerateLogisExpr(self):
        parser = lxml.html.HTMLParser()
        htmltree = xml.etree.ElementTree.fromstring(self.getGoodsDesc(),parser)
        expr = generate_logis_expression_from_html(htmltree)
        print(expr)
        pattern = "[\(\)]" #分离括号
        res = split_logis_expr_and_token(re.split(pattern,expr)[1].strip())
        self.assertEqual(len(res),6)
        self.assertEqual(list(res.keys())[0],"23232445")
        self.assertEqual(list(res.keys())[1],"JT798797947646")
        self.assertEqual(list(res.keys())[2],"232324452-1-2")
        company,notes = list(res.values())[3]
        self.assertEqual(has_special_label(notes),False)
        self.assertEqual(list(res.values())[4],(get_logis_companies()["中通"],["玩具"]))
        company,notes = list(res.values())[5]
        self.assertEqual(has_special_label(notes),True)
        self.assertEqual(len(notes),3)

        manifest = []
        for forecast in get_customer_forecast(htmltree):
            for cnsn,(company,notes) in forecast.items():
                if not has_special_label(notes):
                    manifest.append((cnsn,company,notes))
        expr = generate_manifest_expression(manifest)
        print(expr)
        self.assertEqual(len(manifest),2)
        self.assertEqual(manifest[0][0],"JT7987979476461")
        self.assertEqual(manifest[1][1],get_logis_companies()["中通"])